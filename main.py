import os
import base64
import tempfile
import json
import logging
import time
import threading
from typing import Optional, List
from urllib.parse import urljoin, urlparse
from itertools import product

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from io import BytesIO
from openpyxl import load_workbook

import dropbox
from dropbox.files import WriteMode

import io
import pikepdf

from flask import Flask, request, jsonify, url_for, send_from_directory, render_template_string
from bs4 import BeautifulSoup
from werkzeug.exceptions import HTTPException
from stats_runtime import StatsStore, PROVINCES

from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Flask app & logging
# -----------------------------------------------------------------------------
app = Flask(__name__)

IDEMPOTENCY_TTL_SECONDS = max(60, int(os.getenv("IDEMPOTENCY_TTL_SECONDS", "900")))
_idempotency_lock = threading.Lock()
_idempotency_cache: dict[str, dict[str, object]] = {}


def _idempotency_cleanup_unlocked() -> None:
    now = time.time()
    stale = [k for k, v in _idempotency_cache.items() if float(v.get("expires_at", 0.0)) <= now]
    for k in stale:
        _idempotency_cache.pop(k, None)


def _idempotency_begin(key: Optional[str]) -> tuple[str, Optional[dict[str, object]]]:
    if not key:
        return "run", None
    with _idempotency_lock:
        _idempotency_cleanup_unlocked()
        entry = _idempotency_cache.get(key)
        if not entry:
            _idempotency_cache[key] = {
                "status": "in_progress",
                "expires_at": time.time() + IDEMPOTENCY_TTL_SECONDS,
                "response": None,
            }
            return "run", None
        if entry.get("status") == "done" and isinstance(entry.get("response"), dict):
            return "cached", entry["response"]  # type: ignore[index]
        return "in_progress", None


def _idempotency_finish(key: Optional[str], payload: dict[str, object], status_code: int) -> None:
    if not key:
        return
    with _idempotency_lock:
        if status_code >= 500:
            _idempotency_cache.pop(key, None)
            return
        _idempotency_cache[key] = {
            "status": "done",
            "expires_at": time.time() + IDEMPOTENCY_TTL_SECONDS,
            "response": {"payload": payload, "status_code": int(status_code)},
        }


def _idempotency_abort(key: Optional[str]) -> None:
    if not key:
        return
    with _idempotency_lock:
        entry = _idempotency_cache.get(key)
        if entry and entry.get("status") == "in_progress":
            _idempotency_cache.pop(key, None)


def _idempotency_key(prefix: str) -> Optional[str]:
    value = str(request.headers.get("X-Idempotency-Key", "")).strip()
    if not value:
        return None
    return f"{prefix}:{value[:200]}"

def dropbox_download_file(dropbox_path, token):
    url = 'https://content.dropboxapi.com/2/files/download'
    headers = {
        'Authorization': f'Bearer {token}',
        'Dropbox-API-Arg': json.dumps({'path': dropbox_path})
    }

    resp = requests.post(url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f'Dropbox download error {resp.status_code}: {resp.text}')

    return resp.content


def dropbox_upload_file(dropbox_path, file_bytes, token):
    url = 'https://content.dropboxapi.com/2/files/upload'
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/octet-stream',
        'Dropbox-API-Arg': json.dumps({
            'path': dropbox_path,
            'mode': 'overwrite',
            'autorename': False,
            'mute': True,
            'strict_conflict': False
        })
    }

    resp = requests.post(url, headers=headers, data=file_bytes)
    if resp.status_code != 200:
        raise Exception(f'Dropbox upload error {resp.status_code}: {resp.text}')

    return resp.json()


def safe_sheet_name(name, fallback):
    name = str(name or '').strip()
    if not name:
        name = fallback

    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for ch in invalid_chars:
        name = name.replace(ch, '_')

    return name[:31]

def find_column_by_header(ws, header_name, header_row=1):
    for cell in ws[header_row]:
        if str(cell.value or '').strip().lower() == header_name.strip().lower():
            return cell.column
    return None


def add_dropdown_to_column(ws, header_name, formula1, header_row=1, start_row=2, end_row=5000):
    col_idx = find_column_by_header(ws, header_name, header_row)
    if not col_idx:
        return False

    col_letter = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.error = "Please select a value from the list."
    dv.errorTitle = "Invalid value"
    dv.prompt = f"Choose {header_name} from the dropdown list."
    dv.promptTitle = header_name
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    return True

def write_value_by_header(ws, header_name, value, header_row=1, target_row=2):
    col_idx = find_column_by_header(ws, header_name, header_row)
    if not col_idx:
        return False

    ws.cell(row=target_row, column=col_idx).value = value
    return True

logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

# -----------------------------------------------------------------------------
# HTTP session with timeouts and retries
# -----------------------------------------------------------------------------
DEFAULT_TIMEOUT = 30

def _requests_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": "AR-server/1.1"})
    retries = Retry(
        total=3,
        backoff_factor=0.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET", "POST", "HEAD", "OPTIONS"])
    )
    s.mount("http://", HTTPAdapter(max_retries=retries))
    s.mount("https://", HTTPAdapter(max_retries=retries))
    return s

session = _requests_session()

# -----------------------------------------------------------------------------
# Utility routes
# -----------------------------------------------------------------------------
@app.route("/healthz")
def healthz():
    return "ok", 200

@app.route("/favicon.ico")
def favicon():
    static_path = os.path.join(app.root_path, "static")
    fav = os.path.join(static_path, "favicon.png")
    if os.path.exists(fav):
        return send_from_directory(static_path, "favicon.png", mimetype="image/png")
    return "", 204

# -----------------------------------------------------------------------------
# Main page
# -----------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template_string("""
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>Kenorland Digitizing</title>
  <link rel="icon" href="/static/favicon.png">
  <style>
    :root{
      --bg:#f5f7fa; --fg:#111827; --muted:#6b7280; --card:#ffffff; --sub:#fafafa; --border:#e5e7eb; --accent:#2563eb; --ok:#10b981; --danger:#ef4444;
      --shadow:0 8px 28px rgba(0,0,0,.08);
    }
    *{box-sizing:border-box}
    html,body{min-height:100%}
    body{
      margin:0;
      background:var(--bg);
      color:var(--fg);
      font:15px/1.6 ui-sans-serif,system-ui,-apple-system,Segoe UI,Roboto,"Helvetica Neue",Arial;
      display:flex;
      justify-content:center;
      align-items:flex-start;
      min-height:100vh;
      padding:clamp(10px, 2vw, 24px);
      padding-bottom:clamp(28px, 4vw, 44px);
      overflow-x:hidden;
      overflow-y:auto;
    }

    .wrap{
      width:100%;
      max-width:1200px;
      margin:0 auto;
      background:var(--card);
      border:1px solid var(--border);
      border-radius:16px;
      box-shadow:var(--shadow);
      padding:24px;
    }
    
    header{
      display:flex;
      gap:12px;
      align-items:center;
      margin-bottom:8px;
      flex-wrap:wrap;
    }
    .logo{width:36px;height:36px;border-radius:10px;background:linear-gradient(135deg,#2563eb,#10b981);display:grid;place-items:center;color:#fff;font-weight:700}
    h1{margin:0;font-size:clamp(22px,3vw,30px)}
    .tag{color:var(--ok);font-weight:600;margin-left:auto}
    h2{margin:10px 0 12px;font-size:14px;letter-spacing:.12em;color:var(--muted)}

    /* Two columns */
    .cols{display:grid;grid-template-columns:1fr;gap:16px}
    @media (min-width:980px){ .cols{grid-template-columns:1fr 1fr} } /* balanced columns */

    @media (max-width:768px){
      body{
        padding:12px;
      }

      .wrap{
        padding:16px;
        border-radius:12px;
      }

      h1{
        font-size:24px;
        line-height:1.3;
      }

      .tag{
        margin-left:0;
      }

      section{
        padding:14px;
      }

      .actions{
        justify-content:center;
      }

      .chip,
      .btn{
        font-size:12px;
        padding:8px 10px;
      }

      footer{
        flex-direction:row;
        align-items:center;
        justify-content:space-between;
        flex-wrap:wrap;
        gap:8px;
      }
    }

    @media (max-width:480px){
      body{
        padding:10px;
      }

      .wrap{
        padding:14px;
      }

      h1{
        font-size:21px;
      }

      h2{
        font-size:13px;
      }

      .logo{
        width:32px;
        height:32px;
        border-radius:8px;
      }

      ul{
        margin-left:16px;
      }

      .actions{
        gap:6px;
      }
    }
    
    section{background:var(--sub);border:1px solid var(--border);border-radius:12px;padding:clamp(12px, 1.8vw, 18px)}
    section h3{margin:0 0 6px}
    ul{margin:8px 0 0 18px;padding:0}

    /* Chips and action buttons */
    .actions{display:flex;flex-wrap:wrap;gap:8px;justify-content:center;margin-top:16px}
    .chip{
      display:inline-block;background:#f3f4f6;border:1px solid var(--border);color:#374151;
      font-size:12.5px;line-height:1;padding:7px 12px;border-radius:999px;font-weight:500;
      user-select:none;pointer-events:none; /* read-only */
    }
    .btn{
      border:1px solid var(--border);background:#fff;color:#fg;
      font-size:13px;padding:8px 14px;border-radius:999px;font-weight:700;cursor:pointer;
    }
    .btn:hover{border-color:var(--accent)}
    .btn:active{transform:translateY(.5px)}

    /* modal */
    .backdrop{position:fixed;inset:0;background:rgba(0,0,0,.35);display:none;align-items:center;justify-content:center;padding:16px;z-index:50}
    .modal{width:min(520px,100%);background:#fff;border:1px solid var(--border);border-radius:14px;box-shadow:var(--shadow);padding:18px}
    .modal h4{margin:0 0 8px;font-size:16px}
    .pill{display:inline-block;padding:4px 10px;border-radius:999px;border:1px solid var(--border);font-weight:700}
    .ok{color:var(--ok);border-color:var(--ok)} .bad{color:var(--danger);border-color:var(--danger)}

    footer{display:flex;justify-content:space-between;align-items:center;margin-top:14px;color:var(--muted);font-size:12px}
    footer b{font-weight:700}
    .chart-wrap{
      position:relative;
      width:100%;
      height:clamp(240px, 42vh, 380px);
      min-height:240px;
    }
    #projChart{
      width:100% !important;
      height:100% !important;
      display:block;
    }
    .stats-meta{margin-top:10px;color:#374151;font-size:13px}
    .stats-head{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;flex-wrap:wrap}
    .stats-head-right{display:flex;flex-direction:column;align-items:flex-end;gap:8px}
    .stats-totals{display:flex;flex-wrap:wrap;gap:8px;justify-content:flex-end}
    .stats-box{background:#fff;border:1px solid var(--border);border-radius:10px;padding:6px 10px;font-size:12px}
    .stats-controls{display:flex;flex-wrap:wrap;gap:6px;justify-content:flex-end}
    .period-btn,.refresh-btn{
      border:1px solid var(--border);
      background:#fff;
      color:#1f2937;
      border-radius:999px;
      padding:6px 10px;
      font-size:12px;
      font-weight:600;
      cursor:pointer;
    }
    .period-btn.active{
      border-color:#0057B7;
      color:#0057B7;
      background:#eef5ff;
    }
    .refresh-btn{
      border-color:#FFD700;
      background:#fffdf0;
    }
    .stats-kpis{
      margin-top:10px;
      display:grid;
      grid-template-columns:repeat(4,minmax(120px,1fr));
      gap:8px;
    }
    .kpi-card{
      background:#fff;
      border:1px solid var(--border);
      border-radius:10px;
      padding:8px 10px;
    }
    .kpi-label{font-size:11px;color:#6b7280;text-transform:uppercase;letter-spacing:.06em}
    .kpi-value{font-size:16px;font-weight:700;color:#111827}
    .stats-errors{
      margin-top:10px;
      background:#fff;
      border:1px solid #fecaca;
      border-radius:10px;
      padding:8px 10px;
      font-size:12px;
      color:#7f1d1d;
    }
    #asxChart{width:100% !important;height:100% !important;display:block}
    @media (max-width:768px){
      .stats-head{flex-direction:column;align-items:flex-start}
      .stats-head-right{align-items:flex-start}
      .stats-totals{justify-content:flex-start}
      .stats-controls{justify-content:flex-start}
      .stats-kpis{grid-template-columns:repeat(2,minmax(120px,1fr))}
    }
  </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
</head>
<body>
  <div class="wrap">
    <header>
      <div class="logo">KLD</div>
      <h1>Kenorland Digitizing Server is running 🚀</h1>
      <div class="tag">healthy</div>
    </header>

    <h2>FUNCTIONALITY</h2>
    <div class="cols">
      <section>
        <h3>AR</h3>
        <ul>
          <li>Download AR PDFs for <b>Quebec</b>, <b>Ontario</b>, <b>Manitoba</b></li>
          <li>Create report structure & templates for <b>QC, ON, NB, MB, NU</b>:
            <ul>
              <li>Copy & rename <i>Instructions.xlsx</i></li>
              <li>Copy & rename <i>Geochemistry.gdb</i></li>
              <li>Copy & rename <i>DDH.gdb</i></li>
            </ul>
          </li>
        </ul>
      </section>

      <section>
        <h3>ASX</h3>
        <ul>
          <li>Auto-unlock ASX PDFs</li>
          <li>Removes encryption and copy/print restrictions</li>
          <li>Creates report Excel files from template</li>
          <li>Automatically renames Excel sheets with PDF_ID</li>
          <li>Restores dropdown lists</li>
          <li>Writes PDF_ID into Excel column</li>
        </ul>
      </section>
    </div>

    <div class="actions">
      <span class="chip">Dropbox integrated</span>
      <span class="chip">Google Apps Script integrated</span>
      <span class="chip">Timeouts & retries</span>
      <button class="btn" onclick="checkHealth()">Check health</button>
      <span class="chip">ASX unlock API</span>
    </div>

        <!-- ===== STATS SECTION ===== -->
    <section class="bg-white rounded-2xl shadow p-6" style="margin-top:16px; overflow-x:auto;">
      <div class="stats-head">
        <h2 class="text-xl font-semibold" style="margin:0;">AR Statistics</h2>
        <div class="stats-head-right">
          <div id="statsTotals" class="stats-totals"></div>
          <div class="stats-controls">
            <button class="period-btn" data-period="today">Today</button>
            <button class="period-btn" data-period="7d">7D</button>
            <button class="period-btn" data-period="30d">30D</button>
            <button class="period-btn active" data-period="all">All</button>
            <button class="refresh-btn" id="refreshStatsBtn" type="button">Refresh</button>
          </div>
        </div>
      </div>
      <div id="statsKpis" class="stats-kpis"></div>
      <div class="mt-4">
        <div class="chart-wrap"><canvas id="projChart"></canvas></div>
      </div>
      <div id="statsMeta" class="stats-meta">Tracking start: n/a</div>
      <div id="statsErrors" class="stats-errors" style="display:none"></div>
    </section>

    <section class="bg-white rounded-2xl shadow p-6" style="margin-top:16px; overflow-x:auto;">
      <div class="stats-head">
        <h2 class="text-xl font-semibold" style="margin:0;">ASX Statistics</h2>
        <div id="asxTotals" class="stats-totals"></div>
      </div>
      <div id="asxKpis" class="stats-kpis"></div>
      <div class="mt-4">
        <div class="chart-wrap"><canvas id="asxChart"></canvas></div>
      </div>
      <div id="asxMeta" class="stats-meta">ASX stats period: ALL</div>
      <div id="asxErrors" class="stats-errors" style="display:none"></div>
    </section>

    <footer>
      <div>Powered by <b>Flask</b> &middot; <b>Render</b></div>
      <div>Created by <b>Zirka</b> &middot; <b>chatGPT</b></div> &middot; <b>Codex</b></div>
    </footer>
  </div>

  <!-- Health modal -->
  <div id="backdrop" class="backdrop" role="dialog" aria-modal="true" aria-labelledby="healthTitle">
    <div class="modal">
      <h4 id="healthTitle">Service health</h4>
      <div id="healthBody">Checking...</div>
      <div style="margin-top:12px;display:flex;justify-content:flex-end">
        <button class="btn" onclick="closeModal()">Close</button>
      </div>
    </div>
  </div>

  <script>
    const backdrop = document.getElementById('backdrop');
    const bodyEl = document.getElementById('healthBody');

    function openModal(){ backdrop.style.display = 'flex'; }
    function closeModal(){ backdrop.style.display = 'none'; }

    async function checkHealth(){
      openModal();
      bodyEl.innerHTML = 'Checking...';
      try{
        const res = await fetch('/healthz', { cache:'no-store' });
        const txt = (await res.text() || '').trim();
        const ok = res.ok && txt.toLowerCase().includes('ok');
        bodyEl.innerHTML = ok
          ? 'Status: <span class="pill ok">OK</span>'
          : 'Status: <span class="pill bad">Unavailable</span><div style="margin-top:6px;color:#6b7280">Response: <code>'+escapeHtml(txt)+'</code></div>';
      }catch(e){
        bodyEl.innerHTML = 'Status: <span class="pill bad">Error</span><div style="margin-top:6px;"><code>'+escapeHtml(String(e))+'</code></div>';
      }
    }
    function escapeHtml(s){ return s.replace(/[&<>"']/g, m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m])); }
    backdrop.addEventListener('click', e=>{ if(e.target===backdrop) closeModal(); });
    document.addEventListener('keydown', e=>{ if(e.key==='Escape') closeModal(); });

    // ====== Chart: project stats ======
  let statsChart = null;
  let asxChart = null;
  let currentPeriod = 'all';

  function setActivePeriodBtn(period){
    document.querySelectorAll('.period-btn').forEach(btn=>{
      btn.classList.toggle('active', btn.dataset.period === period);
    });
  }

  async function initChart(period = null) {
    try {
      if (period) currentPeriod = period;
      const res = await fetch('/api/stats?period=' + encodeURIComponent(currentPeriod), { cache: 'no-store' });
      const data = await res.json();
      const labels = data.labels || [];
      const values = data.values || [];
      const templateValues = data.templates_values || [];

      const ctx = document.getElementById('projChart');
      if (!ctx) return;
      if (statsChart) {
        statsChart.destroy();
      }
      const asxCtx = document.getElementById('asxChart');
      if (asxChart) {
        asxChart.destroy();
      }

      statsChart = new Chart(ctx, {
        type: 'bar',
        data: {
          labels,
          datasets: [
            {
              label: 'PDF downloaded',
              data: values,
              backgroundColor: '#FFD700',
              borderColor: '#FFD700',
              borderWidth: 1
            },
            {
              label: 'Folder created',
              data: templateValues,
              backgroundColor: '#0057B7',
              borderColor: '#0057B7',
              borderWidth: 1
            }
          ]
        },
        options: {
          responsive: true,
          maintainAspectRatio: false,
          plugins: { legend: { display: true } },
          scales: { y: { beginAtZero: true, ticks: { precision: 0, stepSize: 1 } } }
        }
      });
      setActivePeriodBtn(data.period || currentPeriod);
      const meta = document.getElementById('statsMeta');
      const fmt = (iso) => {
        if (!iso) return 'n/a';
        const d = new Date(iso);
        if (Number.isNaN(d.getTime())) return iso;
        return d.toLocaleString('en-CA', {
          year: 'numeric',
          month: 'short',
          day: '2-digit',
          hour: '2-digit',
          minute: '2-digit',
          second: '2-digit',
          timeZoneName: 'short'
        });
      };
      if (meta) {
        meta.innerHTML =
          '<b>AR tracking started:</b> ' + fmt(data.tracking_started_at) +
          ' &nbsp;&nbsp;|&nbsp;&nbsp; ' +
          '<b>Last updated:</b> ' + fmt(data.updated_at) +
          ' &nbsp;&nbsp;|&nbsp;&nbsp; ' +
          '<b>Period:</b> ' + String(data.period || currentPeriod).toUpperCase();
      }
      const totals = document.getElementById('statsTotals');
      if (totals && data.totals) {
        totals.innerHTML =
          '<span class="stats-box">PDF total: <b>' + (data.totals.reports_downloaded || 0) + '</b></span>' +
          '<span class="stats-box">Folders total (reports): <b>' + (data.totals.templates_copied || 0) + '</b></span>' +
          '<span class="stats-box">Requests: <b>' + (data.totals.requests || 0) + '</b></span>' +
          '<span class="stats-box">Failed: <b>' + (data.totals.failed_requests || 0) + '</b></span>';
      }

      const kpis = document.getElementById('statsKpis');
      if (kpis && data.kpis) {
        kpis.innerHTML =
          '<div class=\"kpi-card\"><div class=\"kpi-label\">Success rate</div><div class=\"kpi-value\">' + (data.kpis.success_rate ?? 0) + '%</div></div>' +
          '<div class=\"kpi-card\"><div class=\"kpi-label\">Top province</div><div class=\"kpi-value\">' + (data.kpis.top_province || '-') + '</div></div>' +
          '<div class=\"kpi-card\"><div class=\"kpi-label\">Last activity province</div><div class=\"kpi-value\">' + (data.kpis.last_activity_province || '-') + '</div></div>' +
          '<div class=\"kpi-card\"><div class=\"kpi-label\">Last activity</div><div class=\"kpi-value\" style=\"font-size:13px\">' + fmt(data.kpis.last_activity_at) + '</div></div>';
      }

      const errors = document.getElementById('statsErrors');
      if (errors) {
        const recentErrors = Array.isArray(data.recent_errors) ? data.recent_errors : [];
        if (recentErrors.length === 0) {
          errors.style.display = 'none';
          errors.innerHTML = '';
        } else {
          errors.style.display = 'block';
          errors.innerHTML = '<b>Recent errors:</b> ' + recentErrors.map(e => {
            return '[' + fmt(e.ts) + '] ' + e.province;
          }).join(' | ');
        }
      }

      const asx = data.asx || {};
      if (asxCtx) {
        asxChart = new Chart(asxCtx, {
          type: 'bar',
          data: {
            labels: asx.labels || [],
            datasets: [
              {
                label: 'Unlocked & uploaded PDFs',
                data: asx.unlock_values || [],
                backgroundColor: '#0057B7',
                borderColor: '#0057B7',
                borderWidth: 1
              },
              {
                label: 'Created XLSX files',
                data: asx.xlsx_values || [],
                backgroundColor: '#FFD700',
                borderColor: '#FFD700',
                borderWidth: 1
              }
            ]
          },
          options: {
            responsive: true,
            maintainAspectRatio: false,
            plugins: { legend: { display: true } },
            scales: { y: { beginAtZero: true, ticks: { precision: 0, stepSize: 1 } } }
          }
        });
      }

      const asxTotals = document.getElementById('asxTotals');
      if (asxTotals && asx.totals) {
        asxTotals.innerHTML =
          '<span class="stats-box">Unlock requests: <b>' + (asx.totals.unlock_requests || 0) + '</b></span>' +
          '<span class="stats-box">Unlocked PDFs: <b>' + (asx.totals.unlock_uploaded || 0) + '</b></span>' +
          '<span class="stats-box">XLSX requests: <b>' + (asx.totals.xlsx_requests || 0) + '</b></span>' +
          '<span class="stats-box">XLSX created: <b>' + (asx.totals.xlsx_created || 0) + '</b></span>' +
          '<span class="stats-box">Failed: <b>' + (asx.totals.failed_requests || 0) + '</b></span>';
      }

      const asxKpis = document.getElementById('asxKpis');
      if (asxKpis && asx.kpis) {
        asxKpis.innerHTML =
          '<div class="kpi-card"><div class="kpi-label">Success rate</div><div class="kpi-value">' + (asx.kpis.success_rate ?? 0) + '%</div></div>' +
          '<div class="kpi-card"><div class="kpi-label">Last action</div><div class="kpi-value">' + (asx.kpis.last_action || '-') + '</div></div>' +
          '<div class="kpi-card"><div class="kpi-label">Last activity</div><div class="kpi-value" style="font-size:13px">' + fmt(asx.kpis.last_activity_at) + '</div></div>' +
          '<div class="kpi-card"><div class="kpi-label">Period</div><div class="kpi-value">' + String(data.period || currentPeriod).toUpperCase() + '</div></div>';
      }

      const asxMeta = document.getElementById('asxMeta');
      if (asxMeta) {
        asxMeta.innerHTML =
          '<b>ASX tracking started:</b> ' + fmt(data.tracking_started_at) +
          ' &nbsp;&nbsp;|&nbsp;&nbsp; ' +
          '<b>Last updated:</b> ' + fmt(data.updated_at);
      }

      const asxErrors = document.getElementById('asxErrors');
      if (asxErrors) {
        const recentAsxErrors = Array.isArray(asx.recent_errors) ? asx.recent_errors : [];
        if (recentAsxErrors.length === 0) {
          asxErrors.style.display = 'none';
          asxErrors.innerHTML = '';
        } else {
          asxErrors.style.display = 'block';
          asxErrors.innerHTML = '<b>Recent ASX errors:</b> ' + recentAsxErrors.map(e => {
            return '[' + fmt(e.ts) + '] ' + e.action;
          }).join(' | ');
        }
      }
    } catch (e) {
      console.error('Chart init error:', e);
    }
  }

  document.addEventListener('DOMContentLoaded', () => {
    initChart('all');
    document.querySelectorAll('.period-btn').forEach(btn => {
      btn.addEventListener('click', () => initChart(btn.dataset.period));
    });
    const refreshBtn = document.getElementById('refreshStatsBtn');
    if (refreshBtn) {
      refreshBtn.addEventListener('click', () => initChart(currentPeriod));
    }
  });
    
  </script>
</body>
</html>
""")



# -----------------------------------------------------------------------------
# Dropbox helpers
# -----------------------------------------------------------------------------
def get_dropbox_access_token() -> str:
    cid = os.getenv("DROPBOX_CLIENT_ID")
    csec = os.getenv("DROPBOX_CLIENT_SECRET")
    rtok = os.getenv("DROPBOX_REFRESH_TOKEN")
    if not all([cid, csec, rtok]):
        raise RuntimeError("Missing Dropbox credentials")
    auth = base64.b64encode(f"{cid}:{csec}".encode()).decode()
    resp = session.post(
        "https://api.dropbox.com/oauth2/token",
        data={"grant_type": "refresh_token", "refresh_token": rtok},
        headers={"Authorization": f"Basic {auth}"},
        timeout=DEFAULT_TIMEOUT,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]

# -----------------------------------------------------------------------------
# Stats storage (JSON in the same folder as this script)
# -----------------------------------------------------------------------------
STATS_BACKEND = os.getenv("STATS_BACKEND", "file").strip().lower()
STATS_DROPBOX_PATH = os.getenv(
    "STATS_DROPBOX_PATH",
    "/Zirka/OCTOGIT/DIGITIZING/GAppS/github/stats/project_stats.json",
)
STATS_LOCAL_PATH = os.getenv(
    "STATS_LOCAL_PATH",
    os.path.abspath(os.path.join(os.path.dirname(__file__), "stats", "project_stats.json")),
)
_stats_store: StatsStore | None = None


def get_stats_store() -> StatsStore:
    global _stats_store
    if _stats_store is None:
        _stats_store = StatsStore(
            backend=STATS_BACKEND,
            dropbox_path=STATS_DROPBOX_PATH,
            local_path=STATS_LOCAL_PATH,
            logger=app.logger,
            token_provider=get_dropbox_access_token,
        )
    return _stats_store


def track_download_stats(province: str, downloaded_pdfs: int, templates_copied: int, success: bool) -> None:
    if province not in PROVINCES:
        return
    try:
        get_stats_store().apply_download_event(
            province=province,
            downloaded_pdfs=downloaded_pdfs,
            templates_copied=templates_copied,
            success=success,
        )
    except Exception as e:
        app.logger.warning(f"Stats update failed: {e}")


def track_asx_stats(action: str, count: int, success: bool) -> None:
    try:
        get_stats_store().apply_asx_event(action=action, count=count, success=success)
    except Exception as e:
        app.logger.warning(f"ASX stats update failed: {e}")

def ensure_folder(dbx: dropbox.Dropbox, path: str) -> None:
    """Create folder if needed; ignore conflict when already created concurrently."""
    try:
        dbx.files_get_metadata(path)
        return
    except dropbox.exceptions.ApiError:
        pass

    try:
        dbx.files_create_folder_v2(path)
    except dropbox.exceptions.ApiError as e:
        if "conflict" not in str(e).lower() and "folder" not in str(e).lower():
            raise

# -----------------------------------------------------------------------------
# PDF helpers
# -----------------------------------------------------------------------------
def _extract_pdf_links(html: str, base: str) -> List[str]:
    soup = BeautifulSoup(html, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.lower().endswith(".pdf") and urlparse(href).scheme in ("http", "https"):
            links.append(href)
        elif href.lower().endswith(".pdf"):
            links.append(urljoin(base, href))
    return list(dict.fromkeys(links))

def _case_variants(ext: str) -> List[str]:
    if not ext:
        return []
    return [''.join(p) for p in product(*[(c.lower(), c.upper()) for c in ext])]

def _try_get(url: str) -> Optional[bytes]:
    try:
        r = session.get(url, timeout=DEFAULT_TIMEOUT)
        r.raise_for_status()
        return r.content
    except requests.HTTPError:
        return None

def download_ar_generic(ar_number: str, province: str, project: str,
                        list_page_url: str | None = None,
                        base_url: str | None = None,
                        stats_out: dict | None = None) -> int:
    token = get_dropbox_access_token()
    dbx = dropbox.Dropbox(token)
    base = f"/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/1 - NEW REPORTS/{province}/{project}/{ar_number}"
    instr = f"{base}/Instructions"
    srcdata = f"{base}/Source Data"
    for p in (base, instr, srcdata):
        ensure_folder(dbx, p)
    template_copies = 0
    try:
        dbx.files_copy_v2("/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/_Documents/Instructions/01_Instructions.xlsx",
                          f"{instr}/{ar_number}_Instructions.xlsx", autorename=False)
        template_copies += 1
    except dropbox.exceptions.ApiError as e:
        app.logger.warning(f"Instructions copy failed: {e}")
    try:
        dbx.files_copy_v2("/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/_Documents/Instructions/ReportID_Geochemistry.gdb",
                          f"{base}/{ar_number}_Geochemistry.gdb", autorename=False)
        template_copies += 1
    except dropbox.exceptions.ApiError as e:
        app.logger.warning(f"Geochemistry copy failed: {e}")
    try:
        dbx.files_copy_v2("/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/_Documents/Instructions/ReportID_DDH.gdb",
                          f"{base}/{ar_number}_DDH.gdb", autorename=False)
        template_copies += 1
    except dropbox.exceptions.ApiError as e:
        app.logger.warning(f"DDH copy failed: {e}")
    if isinstance(stats_out, dict):
        stats_out["templates_copied"] = 1
    if not list_page_url:
        return 0
    resp = session.get(list_page_url, timeout=DEFAULT_TIMEOUT)
    resp.raise_for_status()
    pdf_links = _extract_pdf_links(resp.text, list_page_url)
    more_links: List[str] = []
    if base_url:
        soup = BeautifulSoup(resp.text, "html.parser")
        hrefs = [a["href"].strip() for a in soup.find_all("a", href=True)]
        candidates = []
        for h in hrefs:
            name = os.path.basename(h)
            root, ext = os.path.splitext(name)
            if ext and ext[1:].lower() == "pdf":
                candidates.append(root)
        if not candidates:
            for h in hrefs:
                name = os.path.basename(h)
                root, _ = os.path.splitext(name)
                if root:
                    candidates.append(root)
        candidates = list(dict.fromkeys(candidates))
        for root in candidates:
            for v in _case_variants("pdf"):
                more_links.append(f"{base_url}/{ar_number}/{root}.{v}")
    all_links = list(dict.fromkeys(pdf_links + more_links))
    count = 0
    for url in all_links:
        try:
            content = _try_get(url)
            if not content:
                continue
            filename = os.path.basename(urlparse(url).path) or "file.pdf"
            dbx.files_upload(content, f"{srcdata}/{filename}", mode=WriteMode.overwrite)
            count += 1
        except Exception as e:
            app.logger.error(f"PDF upload error [{url}]: {e}")
    return count


# -------------------- ADD: Manitoba direct-download logic --------------------
def download_ar_manitoba(ar_number: str, province: str, project: str, stats_out: dict | None = None) -> int:
    """
    Manitoba: direct download of a single PDF:
    https://www.gov.mb.ca/data/em/application/assessment/{ar_number}.pdf
    """
    token = get_dropbox_access_token()
    dbx = dropbox.Dropbox(token)

    base = f"/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/1 - NEW REPORTS/{province}/{project}/{ar_number}"
    instr = f"{base}/Instructions"
    srcdata = f"{base}/Source Data"

    # Create folders and copy templates (same flow as other provinces)
    for p in (base, instr, srcdata):
        ensure_folder(dbx, p)
    template_copies = 0
    try:
        dbx.files_copy_v2("/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/_Documents/Instructions/01_Instructions.xlsx",
                          f"{instr}/{ar_number}_Instructions.xlsx", autorename=False)
        template_copies += 1
    except dropbox.exceptions.ApiError as e:
        app.logger.warning(f"Instructions copy failed: {e}")
    try:
        dbx.files_copy_v2("/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/_Documents/Instructions/ReportID_Geochemistry.gdb",
                          f"{base}/{ar_number}_Geochemistry.gdb", autorename=False)
        template_copies += 1
    except dropbox.exceptions.ApiError as e:
        app.logger.warning(f"Geochemistry copy failed: {e}")
    try:
        dbx.files_copy_v2("/KENORLAND_DIGITIZING/ASSESSMENT_REPORTS/_Documents/Instructions/ReportID_DDH.gdb",
                          f"{base}/{ar_number}_DDH.gdb", autorename=False)
        template_copies += 1
    except dropbox.exceptions.ApiError as e:
        app.logger.warning(f"DDH copy failed: {e}")
    if isinstance(stats_out, dict):
        stats_out["templates_copied"] = 1

    # Direct PDF URL without suffix permutations
    url = f"https://www.gov.mb.ca/data/em/application/assessment/{ar_number}.pdf"
    content = _try_get(url)
    if not content:
        return 0

    filename = os.path.basename(urlparse(url).path) or f"{ar_number}.pdf"
    dbx.files_upload(content, f"{srcdata}/{filename}", mode=WriteMode.overwrite)
    return 1
# -----------------------------------------------------------------------------


ASX_DROPBOX_PREFIX = "/KENORLAND_DIGITIZING/ASX/2 - WORKING/"

def _is_allowed_asx_path(path: str) -> bool:
    return isinstance(path, str) and path.startswith(ASX_DROPBOX_PREFIX)

def _check_bearer(req) -> None:
    expected = os.getenv("ASX_UNLOCK_TOKEN", "").strip()
    if not expected:
        return
    if req.headers.get("Authorization", "") != f"Bearer {expected}":
        raise PermissionError("Unauthorized")

def _unlock_pdf_bytes(data: bytes) -> bytes:
    """
    Removes owner restrictions and writes an unencrypted copy when possible.
    Supports files that can be opened with an empty user password ("").
    If a real non-empty user password is required or any error happens, return original bytes.
    """
    import io
    import pikepdf

    # Try without password first, then with an empty password
    for pw in (None, ""):
        try:
            pdf = pikepdf.open(io.BytesIO(data), password=pw)
            try:
                out = io.BytesIO()
                # IMPORTANT: save without encryption args to produce a plain, non-password PDF
                pdf.save(out)
                pdf.close()
                return out.getvalue()
            finally:
                try:
                    pdf.close()
                except Exception:
                    pass
        except (pikepdf.PasswordError, getattr(pikepdf, "_qpdf", type("", (), {})) .__dict__.get("PasswordError", Exception)):
            # Real user password is required; keep original file
            continue
        except Exception:
            # Any other error: fall back to original bytes
            break

    return data


@app.post("/asx_unlock_upload")
def asx_unlock_upload():
    idem_key = _idempotency_key("asx_unlock_upload")
    idem_state, idem_cached = _idempotency_begin(idem_key)
    if idem_state == "cached" and isinstance(idem_cached, dict):
        return jsonify(idem_cached.get("payload", {})), int(idem_cached.get("status_code", 200))
    if idem_state == "in_progress":
        return jsonify(error="Duplicate request in progress"), 409

    try:
        _check_bearer(request)
    except PermissionError:
        payload = {"error": "Unauthorized"}
        _idempotency_finish(idem_key, payload, 401)
        return jsonify(payload), 401

    f = request.files.get("file")
    path = request.form.get("dropbox_path", "")
    if not f or not path:
        payload = {"error": "Missing file or dropbox_path"}
        _idempotency_finish(idem_key, payload, 400)
        return jsonify(payload), 400
    if not _is_allowed_asx_path(path):
        payload = {"error": "Path not allowed"}
        _idempotency_finish(idem_key, payload, 400)
        return jsonify(payload), 400

    try:
        data = f.read()
        if not data:
            payload = {"error": "Empty file"}
            _idempotency_finish(idem_key, payload, 400)
            return jsonify(payload), 400

        unlocked = _unlock_pdf_bytes(data)

        token = get_dropbox_access_token()
        dbx = dropbox.Dropbox(token)
        dbx.files_upload(unlocked, path, mode=WriteMode.overwrite)
        track_asx_stats("unlock_upload", 1, True)

        payload = {"message": "Uploaded (unlocked if possible)", "path": path}
        _idempotency_finish(idem_key, payload, 200)
        return jsonify(payload), 200
    except Exception as e:
        track_asx_stats("unlock_upload", 0, False)
        app.logger.error(f"/asx_unlock_upload error: {e}", exc_info=True)
        _idempotency_abort(idem_key)
        return jsonify(error=str(e)), 500


# -----------------------------------------------------------------------------
# API route
# -----------------------------------------------------------------------------
@app.route("/download_gm", methods=["POST"])
def download_gm():
    idem_key = _idempotency_key("download_gm")
    idem_state, idem_cached = _idempotency_begin(idem_key)
    if idem_state == "cached" and isinstance(idem_cached, dict):
        return jsonify(idem_cached.get("payload", {})), int(idem_cached.get("status_code", 200))
    if idem_state == "in_progress":
        return jsonify(error="Duplicate request in progress"), 409

    data = request.get_json(force=True, silent=True) or {}
    num  = str(data.get("ar_number", "")).strip()
    prov = str(data.get("province", "")).strip()
    proj = str(data.get("project", "")).strip()
    if not all([num, prov, proj]):
        payload = {"error": "Missing parameters"}
        _idempotency_finish(idem_key, payload, 400)
        return jsonify(payload), 400
    cnt = 0
    stats_out: dict = {}
    tpl = 0
    try:
        if prov == "Quebec" and num.upper().startswith("GM"):
            url = f"https://gq.mines.gouv.qc.ca/documents/EXAMINE/{num}/"
            cnt = download_ar_generic(num, prov, proj, url, stats_out=stats_out)
        elif prov == "Ontario":
            url = f"https://www.geologyontario.mndm.gov.on.ca/mndmfiles/afri/data/records/{num}.html"
            blob = "https://prd-0420-geoontario-0000-blob-cge0eud7azhvfsf7.z01.azurefd.net/lrc-geology-documents/assessment"
            cnt = download_ar_generic(num, prov, proj, url, blob, stats_out=stats_out)
        elif prov == "New Brunswick":
            cnt = download_ar_generic(num, prov, proj, stats_out=stats_out)
        elif prov == "Nunavut":
            cnt = download_ar_generic(num, prov, proj, stats_out=stats_out)
        elif prov == "Manitoba":
            cnt = download_ar_manitoba(num, prov, proj, stats_out=stats_out)
        else:
            payload = {"error": "Invalid province or AR#"}
            _idempotency_finish(idem_key, payload, 400)
            return jsonify(payload), 400
        tpl = int(stats_out.get("templates_copied", 0) or 0)
        track_download_stats(prov, cnt, tpl, True)
        msg = f"Downloaded {cnt} PDFs" if cnt > 0 else "Folders created. No PDFs downloaded."
        payload = {"message": msg, "downloaded_pdfs": cnt, "templates_copied": tpl}
        _idempotency_finish(idem_key, payload, 200)
        return jsonify(payload), 200
    except requests.HTTPError as he:
        track_download_stats(prov, cnt, tpl, False)
        app.logger.error(f"HTTP error: {he}", exc_info=True)
        _idempotency_abort(idem_key)
        return jsonify(error=str(he)), 502
    except Exception as e:
        track_download_stats(prov, cnt, tpl, False)
        app.logger.error(f"Unexpected error: {e}", exc_info=True)
        _idempotency_abort(idem_key)
        return jsonify(error=str(e)), 500

# -----------------------------------------------------------------------------
# Error handler
# -----------------------------------------------------------------------------
@app.errorhandler(Exception)
def all_errors(e):
    if isinstance(e, HTTPException):
        return e
    app.logger.error(f"Unhandled: {e}", exc_info=True)
    return jsonify(error="Internal server error"), 500

# -----------------------------------------------------------------------------
# Local run
# -----------------------------------------------------------------------------

@app.get("/api/stats")
def api_stats():
    try:
        period = str(request.args.get("period", "all")).strip().lower()
        state = get_stats_store().load()
        payload = StatsStore.to_api_payload(state, period=period)
        return jsonify(payload), 200
    except Exception as e:
        app.logger.warning(f"Stats read failed: {e}")
        fallback = StatsStore.to_api_payload(None, period="all")
        return jsonify(fallback), 200

@app.route('/asx_create_xlsx_test', methods=['POST'])
def asx_create_xlsx_test():
    data = request.get_json(silent=True) or {}
    return jsonify({
        "ok": True,
        "message": "Test endpoint works",
        "received": data
    }), 200

@app.route('/asx_create_xlsx_rename_test', methods=['POST'])
def asx_create_xlsx_rename_test():
    data = request.get_json(silent=True) or {}
    report_id = str(data.get('report_id') or '').strip()

    if not report_id:
        return jsonify({
            "ok": False,
            "error": "report_id is required"
        }), 400

    wb = Workbook()

    # The first worksheet is created automatically
    ws1 = wb.active
    ws1.title = 'Report_ID_Drilling'

    # Create the second worksheet manually
    ws2 = wb.create_sheet('Report_ID_SurfaceGeochemistry')

    # Build new worksheet names
    drilling_name = f'{report_id}_Drilling'
    surface_name = f'{report_id}_SurfaceGeochemistry'

    # Excel limits worksheet title length to 31 chars
    if len(drilling_name) > 31:
        drilling_name = drilling_name[:31]

    if len(surface_name) > 31:
        surface_name = surface_name[:31]

    # Rename worksheets
    wb['Report_ID_Drilling'].title = drilling_name
    wb['Report_ID_SurfaceGeochemistry'].title = surface_name

    # Save in memory to validate workbook structure
    output = BytesIO()
    wb.save(output)

    return jsonify({
        "ok": True,
        "message": "XLSX rename test works",
        "report_id": report_id,
        "sheet_names": wb.sheetnames
    }), 200

@app.route('/asx_create_xlsx_dropbox_test', methods=['POST'])
def asx_create_xlsx_dropbox_test():
    idem_key = _idempotency_key("asx_create_xlsx_dropbox_test")
    idem_state, idem_cached = _idempotency_begin(idem_key)
    if idem_state == "cached" and isinstance(idem_cached, dict):
        return jsonify(idem_cached.get("payload", {})), int(idem_cached.get("status_code", 200))
    if idem_state == "in_progress":
        return jsonify(error="Duplicate request in progress"), 409

    data = request.get_json(silent=True) or {}

    report_id = str(data.get('report_id') or '').strip()
    template_path = str(data.get('template_path') or '').strip()
    output_path = str(data.get('output_path') or '').strip()

    if not report_id:
        payload = {"ok": False, "error": "report_id is required"}
        _idempotency_finish(idem_key, payload, 400)
        return jsonify(payload), 400

    if not template_path:
        payload = {"ok": False, "error": "template_path is required"}
        _idempotency_finish(idem_key, payload, 400)
        return jsonify(payload), 400

    if not output_path:
        payload = {"ok": False, "error": "output_path is required"}
        _idempotency_finish(idem_key, payload, 400)
        return jsonify(payload), 400

    try:
        token = get_dropbox_access_token()
    except Exception as e:
        _idempotency_abort(idem_key)
        return jsonify({"ok": False, "error": f"Dropbox auth failed: {str(e)}"}), 500

    try:
        # 1. Download template from Dropbox
        file_bytes = dropbox_download_file(template_path, token)

        # 2. Open workbook from memory
        wb = load_workbook(BytesIO(file_bytes))

        # 3. Build new worksheet names
        drilling_name = safe_sheet_name(f'{report_id}_Drilling', 'Drilling')
        surface_name = safe_sheet_name(f'{report_id}_SurfaceGeochemistry', 'SurfaceGeochemistry')

        # 4. Rename worksheets if present
        renamed = []

        if 'Report_ID_Drilling' in wb.sheetnames:
            ws_drill = wb['Report_ID_Drilling']
            ws_drill.title = drilling_name
            write_value_by_header(ws_drill, 'PDF_ID', report_id)

            add_dropdown_to_column(ws_drill, 'Country', '=Info!$A$2:$A$100')
            add_dropdown_to_column(ws_drill, 'UtmZone', '=Info!$B$2:$B$100')
            add_dropdown_to_column(ws_drill, 'HoleType', '=Info!$C$2:$C$100')
            add_dropdown_to_column(ws_drill, 'HoleSize', '=Info!$J$2:$J$350')
            add_dropdown_to_column(ws_drill, 'SampleType', '=Info!$H$2:$H$100')
            add_dropdown_to_column(ws_drill, 'Sample_Medium', '=Info!$G$2:$G$100')

            renamed.append(drilling_name)

        if 'Report_ID_SurfaceGeochemistry' in wb.sheetnames:
            ws_surface = wb['Report_ID_SurfaceGeochemistry']
            ws_surface.title = surface_name
            write_value_by_header(ws_surface, 'PDF_ID', report_id)

            add_dropdown_to_column(ws_surface, 'Country', '=Info!$A$2:$A$100')
            add_dropdown_to_column(ws_surface, 'UtmZone', '=Info!$B$2:$B$100')
            add_dropdown_to_column(ws_surface, 'HoleType', '=Info!$C$2:$C$100')
            add_dropdown_to_column(ws_surface, 'HoleSize', '=Info!$J$2:$J$350')
            add_dropdown_to_column(ws_surface, 'SampleType', '=Info!$H$2:$H$100')
            add_dropdown_to_column(ws_surface, 'Sample_Medium', '=Info!$G$2:$G$100')

            renamed.append(surface_name)

        # 5. Save workbook in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 6. Upload the updated workbook back to Dropbox
        upload_result = dropbox_upload_file(output_path, output.getvalue(), token)
        track_asx_stats("xlsx_create", 1, True)

        payload = {
            "ok": True,
            "message": "Dropbox XLSX created successfully",
            "report_id": report_id,
            "template_path": template_path,
            "output_path": output_path,
            "sheet_names": wb.sheetnames,
            "renamed": renamed,
            "dropbox_result": upload_result
        }
        _idempotency_finish(idem_key, payload, 200)
        return jsonify(payload), 200

    except Exception as e:
        track_asx_stats("xlsx_create", 0, False)
        _idempotency_abort(idem_key)
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port)


